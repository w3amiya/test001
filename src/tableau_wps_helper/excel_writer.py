from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_to_tuple


SUPPORTED_SOURCE_SUFFIXES = {".csv", ".xlsx", ".xlsm"}


@dataclass(frozen=True)
class ImportResult:
    rows_written: int
    columns_written: int
    target_workbook: str
    target_sheet: str


class ExcelImportError(RuntimeError):
    pass


def _read_csv_rows(path: Path) -> list[list[object]]:
    encodings = ("utf-8-sig", "utf-8", "gb18030")
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return [row for row in csv.reader(handle)]
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ExcelImportError(f"Unable to decode CSV file: {path}") from last_error


def _read_xlsx_rows(path: Path) -> list[list[object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return rows


def read_source_rows(source_path: str | Path) -> list[list[object]]:
    path = Path(source_path).expanduser()
    if not path.exists():
        raise ExcelImportError(f"Source file not found: {path}")
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_SOURCE_SUFFIXES:
        raise ExcelImportError(f"Unsupported source file type: {suffix}")
    if suffix == ".csv":
        return _read_csv_rows(path)
    return _read_xlsx_rows(path)


def _max_width(rows: Iterable[list[object]]) -> int:
    return max((len(row) for row in rows), default=0)


def clear_values_preserve_format(sheet, start_row: int, start_col: int) -> None:
    for row in sheet.iter_rows(
        min_row=start_row,
        min_col=start_col,
        max_row=sheet.max_row,
        max_col=sheet.max_column,
    ):
        for cell in row:
            cell.value = None


def import_file_to_workbook(
    source_path: str | Path,
    workbook_path: str | Path,
    sheet_name: str,
    start_cell: str = "A1",
    header_rows: int = 1,
) -> ImportResult:
    source = Path(source_path).expanduser()
    target = Path(workbook_path).expanduser()
    rows = read_source_rows(source)
    start_row, start_col = coordinate_to_tuple(start_cell)
    if target.exists():
        workbook = load_workbook(target)
    else:
        target.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()

    try:
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
                default_sheet = workbook["Sheet"]
                if default_sheet.max_row == 1 and default_sheet.max_column == 1 and default_sheet["A1"].value is None:
                    workbook.remove(default_sheet)
        else:
            sheet = workbook[sheet_name]
        clear_values_preserve_format(sheet, start_row, start_col)

        for row_offset, row in enumerate(rows):
            for col_offset, value in enumerate(row):
                sheet.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)

        workbook.save(target)
    finally:
        workbook.close()

    return ImportResult(
        rows_written=max(len(rows) - max(header_rows, 0), 0),
        columns_written=_max_width(rows),
        target_workbook=str(target),
        target_sheet=sheet_name,
    )
